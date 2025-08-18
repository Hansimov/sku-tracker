import argparse
import json
import pandas as pd
import openpyxl
import sys
import warnings

from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from tclogger import logger, logstr, brk, match_val, get_date_str, str_to_t
from tclogger import dict_to_str, dict_to_table_str
from typing import Union, Literal
from datetime import datetime, timedelta
from collections import defaultdict

from configs.envs import DATA_ROOT, LOCATION_LIST, LOCATION_MAP
from configs.envs import SKIP_WEBSITE_CHECKS_MAP, WEBSITE_NAMES

warnings.filterwarnings("ignore", category=FutureWarning)

DISCOUNT_COLUMNS_MAP = {
    "blinkit": {
        "disc": "Disc_Blinkit",
        "price": "price_blinkit",
        "mrp": "mrp_blinkit",
    },
    "swiggy": {
        "disc": "Disc_Instamart",
        "price": "price_instamart",
        "mrp": "mrp_instamart",
    },
    "zepto": {
        "disc": "Disc_Zepto",
        "price": "price_zepto",
        "mrp": "mrp_zepto",
    },
    "zepto_supersaver": {
        "disc": "Disc_ZeptoSuperSaver",
        "price": "price_supersaver_zepto",
        "mrp": "mrp_zepto",
    },
    "dmart": {
        "disc": "Disc_Dmart",
        "price": "price_dmart",
        "mrp": "mrp_dmart",
    },
}
EXCLUDE_COLUMNS = [
    "location_blinkit",
    "location_zepto",
    "location_instamart",
    "location_dmart",
]
WEBSITE_CHECK_COLUMNS_MAP = {
    "blinkit": {
        "link": "weblink_blinkit",
        "checks": ["instock_blinkit"],
    },
    "zepto": {
        "link": "weblink_zepto",
        "checks": ["instock_zepto"],
    },
    "swiggy": {
        "link": "weblink_instamart",
        "checks": ["instock_instamart"],
    },
    "dmart": {
        "link": "weblink_dmart",
        "checks": ["instock_dmart"],
    },
}
CHECK_GROUP_KEYS = ["website", "location"]


def get_location_val(location: str) -> str:
    return LOCATION_MAP.get(location, location)


def merge_dfs_horizontally(dfs: list[pd.DataFrame]) -> pd.DataFrame:
    """Merge all dfs into one df: concatenate columns"""
    if not dfs:
        return pd.DataFrame()
    merged_df = dfs[0].copy()
    for df in dfs[1:]:
        for col in df.columns:
            if col not in merged_df.columns:
                merged_df[col] = df[col]
            else:
                # update each row with non-nan value
                mask = merged_df[col].isna() | merged_df[col].eq("")
                merged_df.loc[mask, col] = df.loc[mask, col]
    return merged_df


def merge_dfs_vertically(dfs: list[pd.DataFrame]) -> pd.DataFrame:
    """Merge all dfs into one df: concatenate rows"""
    if not dfs:
        return pd.DataFrame()
    merged_df = pd.concat(dfs, ignore_index=True)
    return merged_df


def merge_dfs(
    dfs: list[pd.DataFrame],
    direction: Literal["horizontal", "vertical"] = "horizontal",
) -> pd.DataFrame:
    if direction.lower().startswith("h"):
        return merge_dfs_horizontally(dfs)
    elif direction.lower().startswith("v"):
        return merge_dfs_vertically(dfs)
    else:
        raise ValueError(f"Invalid merge direction: {direction}")


def log_df_tail(df: pd.DataFrame, n: int = 5):
    logger.mesg(f"> DataFrame tail {n} rows:")
    with pd.option_context("display.show_dimensions", False):
        logger.line(df.tail(n), indent=2)


def log_df_dims(df: pd.DataFrame):
    row_cnt, col_cnt = df.shape
    logger.mesg(f"* [{logstr.file(row_cnt)} rows x {logstr.file(col_cnt)} cols]")


def read_df_from_xlsx(xlsx_path: Path, verbose: bool = True) -> pd.DataFrame:
    """Read all sheets from xlsx file and merge them into one DataFrame"""
    xlsx = pd.ExcelFile(xlsx_path, engine="openpyxl")
    df_list = []
    for sheet_name in xlsx.sheet_names:
        df = pd.read_excel(xlsx_path, sheet_name=sheet_name, engine="openpyxl")
        df_list.append(df)
    merge_df = merge_dfs(df_list, "vertical")
    if verbose:
        log_df_tail(merge_df)
        log_df_dims(merge_df)
    return merge_df


class DataframeEditor:
    def remove_columns(
        self,
        df: pd.DataFrame,
        columns: list[str] = EXCLUDE_COLUMNS,
        inplace: bool = True,
    ) -> pd.DataFrame:
        logger.note(f"> Removing columns: {logstr.file(columns)}")
        if not inplace:
            df = df.copy()
        df_columns = df.columns.tolist()
        for col in columns:
            col_name, _, _ = match_val(col, df_columns, use_fuzz=True)
            if col_name in df.columns:
                df.drop(columns=col_name, inplace=True)
        return df

    def insert_date_and_location_columns(
        self, df: pd.DataFrame, location_name: str, date_str: str
    ) -> pd.DataFrame:
        logger.note(f"> Inserting date and location columns ...")
        if "date" not in df.columns:
            date_val = date_str.replace("-", "/")
            df.insert(0, "Date", date_val)
        if "location" not in df.columns:
            location_val = get_location_val(location_name)
            df.insert(1, "Location", location_val)
        return df

    def check_price(self, price: Union[float, int]) -> Union[bool, float]:
        if not price:
            return False
        if isinstance(price, str):
            try:
                price = float(price)
            except Exception as e:
                raise e
        if isinstance(price, (float, int)):
            if price <= 0:
                return False
            else:
                return price
        else:
            raise ValueError(f"invalid price type: {price}")

    def insert_discount_columns(
        self, df: pd.DataFrame, val_format: Literal["float", "percent"] = "float"
    ) -> pd.DataFrame:
        """discount = (1 - price/mrp) * 100 %"""
        logger.note(f"> Inserting discount columns ...")
        for _, col_map in DISCOUNT_COLUMNS_MAP.items():
            columns = df.columns.tolist()
            discount_col = col_map["disc"]

            price_col_name, price_col_idx, _ = match_val(
                col_map["price"], columns, use_fuzz=True
            )
            mrp_col_name, mrp_col_idx, _ = match_val(
                col_map["mrp"], columns, use_fuzz=True
            )

            if price_col_name is None or mrp_col_name is None:
                continue

            discount_values = []
            for _, row in df.iterrows():
                price = row[price_col_name]
                mrp = row[mrp_col_name]
                if val_format == "percent":
                    discount_val_default = ""
                else:
                    discount_val_default = ""
                try:
                    price_num = self.check_price(price)
                    mrp_num = self.check_price(mrp)
                    if price_num is False or mrp_num is False:
                        discount_values.append(discount_val_default)
                        continue
                    discount = 1 - price_num / mrp_num
                    if val_format == "percent":
                        discount_val = f"{discount*100:.0f}%"
                    else:
                        discount_val = round(discount, 2)
                    discount_values.append(discount_val)
                except Exception as e:
                    discount_values.append(discount_val_default)
                    logger.warn(f"× Cannot calc discount: {e}")

            discount_col_idx = price_col_idx + 1
            df.insert(discount_col_idx, discount_col, discount_values)
        return df


class ExcelMerger:
    def __init__(self, date_str: str = None):
        self.date_str = get_date_str(date_str)
        self.editor = DataframeEditor()
        self.init_paths()
        self.init_workbook()

    def init_paths(self):
        self.output_root = DATA_ROOT / "output" / self.date_str
        self.output_merge_path = self.output_root / f"sku_{self.date_str}.xlsx"

    def init_workbook(self):
        self.workbook = openpyxl.Workbook()
        # Remove default sheet created by openpyxl
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

    def get_xlsx_paths_by_location(self, location_name: str) -> list[Path]:
        """Extract xlsx files end with same location_name for each website"""
        xlsx_paths = []
        for website in WEBSITE_NAMES:
            output_dir_by_website = self.output_root / website
            if not output_dir_by_website.exists():
                logger.warn(f"  × No output for website: {brk(website)}")
                continue
            for xlsx_path in output_dir_by_website.glob(f"*_{location_name}.xlsx"):
                if xlsx_path.is_file():
                    xlsx_paths.append(xlsx_path)
        return xlsx_paths

    def read_df_list_from_xlsx_files_with_same_location(
        self, location_name: str
    ) -> list[pd.DataFrame]:
        """Read all xlsx files with same location name, and return a list of DataFrames"""
        logger.note(
            f"> Reading xlsx files for location: {logstr.mesg(brk(location_name))}"
        )
        xlsx_paths = self.get_xlsx_paths_by_location(location_name)
        if not xlsx_paths:
            raise FileNotFoundError(
                f"No xlsx files found for location: {location_name}"
            )
        df_list = []
        for xlsx_path in xlsx_paths:
            df = pd.read_excel(
                xlsx_path, header=0, engine="openpyxl", keep_default_na=False
            )
            df_list.append(df)
        return df_list

    def set_sheet_styles(self, sheet: Worksheet):
        logger.note(f"> Setting sheet styles ...")
        # set discount columns number_format to percentage
        for _, col_map in DISCOUNT_COLUMNS_MAP.items():
            discount_col = col_map["disc"]
            # get discount column idx in header
            for col_idx, cell in enumerate(sheet[1], 1):
                if cell.value == discount_col:
                    # apply to entire column
                    for row in range(2, sheet.max_row + 1):
                        sheet.cell(row=row, column=col_idx).number_format = "0%"
                    break

    def write_df_to_sheet(self, df: pd.DataFrame, location_name: str):
        """Write dataframe to new sheet in workbook"""
        sheet_name = f"{self.date_str}_{get_location_val(location_name)}"
        sheet = self.workbook.create_sheet(title=sheet_name)
        # write headers
        for col_idx, column in enumerate(df.columns, 1):
            sheet.cell(row=1, column=col_idx, value=column)

        # write data
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        self.set_sheet_styles(sheet)

    def merge(self):
        logger.note(f"> Merging xlsx files for:")
        logger.mesg(f"  * locations: {logstr.file(LOCATION_LIST)}")
        logger.mesg(f"  * websites : {logstr.file(WEBSITE_NAMES)}")
        for location_name in LOCATION_LIST:
            df_list = self.read_df_list_from_xlsx_files_with_same_location(
                location_name
            )
            merged_df = merge_dfs(df_list, "horizontal")
            merged_df = self.editor.insert_discount_columns(merged_df)
            merged_df = self.editor.insert_date_and_location_columns(
                merged_df, location_name, self.date_str
            )
            merged_df = self.editor.remove_columns(merged_df)
            print(merged_df)
            self.write_df_to_sheet(merged_df, location_name)

        logger.note(f"> Save merged xlsx to:")
        self.workbook.save(self.output_merge_path)
        logger.okay(f"  * {self.output_merge_path}")


class ExcelChecker:
    """Check rows in Excel file (per day) for missing or invalid data."""

    def __init__(self, date_str: str = None):
        self.date_str = get_date_str(date_str)
        self.init_paths()

    def init_paths(self):
        self.root = DATA_ROOT / "output" / self.date_str
        self.xlsx_path = self.root / f"sku_{self.date_str}.xlsx"
        self.log_path = self.xlsx_path.with_suffix(".log")

    def get_should_skip_check(
        self,
        website: str,
        link_val: str,
        df_columns: list[str],
        row: pd.Series,
    ) -> bool:
        """skip specific conditions"""
        # skip empty link
        if pd.isna(link_val):
            return True
        # skip invalid cell value
        if website in SKIP_WEBSITE_CHECKS_MAP.keys():
            # if any skip_cond matches, skip check
            # and in each skip_cond, all sub_skip_conds must match
            should_skip_check = False
            for skip_conds in SKIP_WEBSITE_CHECKS_MAP[website]:
                should_skip_check = True
                for skip_col, skip_val in skip_conds.items():
                    skip_col_name, _, _ = match_val(skip_col, df_columns, use_fuzz=True)
                    # if any sub_skip_cond not match,
                    # break, and try to match next skip_cond
                    if skip_val != row.get(skip_col_name, None):
                        should_skip_check = False
                        break
                # meet one matched skip_cond, skip later sub_skip_conds
                if should_skip_check:
                    break
            return should_skip_check
        return False

    def count_issues(
        self, issues: list[dict], res_format: Literal["dict", "list"] = "dict"
    ) -> Union[dict[tuple, int], list[dict]]:
        group_res = defaultdict(list)
        for issue in issues:
            key = tuple(issue[k] for k in CHECK_GROUP_KEYS)
            group_res[key].append(issue)
        if res_format == "list":
            res = [
                dict(zip(CHECK_GROUP_KEYS, k)) | {"num": len(v)}
                for k, v in group_res.items()
            ]
        else:
            res = {k: len(v) for k, v in group_res.items()}
        return res

    def format_check_res(
        self, res: list[dict], output_format: Literal["dict", "table"] = "table"
    ) -> str:
        count_res = self.count_issues(res)
        if output_format == "table":
            return dict_to_table_str(
                count_res,
                key_headers=CHECK_GROUP_KEYS,
                val_headers=["Num"],
                aligns=["r", "r", "r"],
                sum_at_tail=True,
                is_colored=False,
            )
        else:  # output_format == "dict":
            return dict_to_str(count_res)

    def dump_check_res(self, res: list[dict]):
        logger.note(f"> Dump checks log:")
        if not res:
            log_res = {
                "counts": {},
                "issues": [],
            }
        else:
            count_res = self.count_issues(res, res_format="list")
            log_res = {
                "counts": count_res,
                "issues": res,
            }

        if not self.log_path.parent.exists():
            self.log_path.parent.mkdir(parents=True, exist_ok=True)
        with open(self.log_path, "w") as wf:
            json.dump(log_res, wf, indent=4, ensure_ascii=False)
        logger.file(f"  * {self.log_path}")

    def check(self, verbose: bool = False) -> dict:
        """
        Example output:
        [
            {
                "website": "swiggy",
                "location": "...",
                "date": ...,
                "link": "weblink_swiggy",
                "column": "instock_swiggy",
                "value": "N/A",
                "row": ...
            },
            ...
        ]
        """
        logger.note(f"> Checking xlsx file:")
        logger.file(f"  * {self.xlsx_path}")
        if not self.xlsx_path.exists():
            logger.warn(f"  × Excel does not exist!")
            return False
        res = []
        df = read_df_from_xlsx(self.xlsx_path)
        df_columns = df.columns.tolist()
        location_col_name, _, _ = match_val("location", df_columns, use_fuzz=True)
        date_col_name, _, _ = match_val("date", df_columns, use_fuzz=True)
        issue_values = ["", "n/a", None]
        for website, check_cols in WEBSITE_CHECK_COLUMNS_MAP.items():
            link_col_name, _, _ = match_val(
                check_cols["link"], df_columns, use_fuzz=True
            )
            for check_col in check_cols["checks"]:
                check_col_name, _, _ = match_val(check_col, df_columns, use_fuzz=True)
                for row_idx, row in df.iterrows():
                    link_val = row.get(link_col_name, None)
                    if self.get_should_skip_check(
                        website=website,
                        df_columns=df_columns,
                        link_val=link_val,
                        row=row,
                    ):
                        continue
                    cell_val = row.get(check_col_name, None)
                    if pd.isna(cell_val) or cell_val in issue_values:
                        issue_item = {
                            "website": website,
                            "location": row.get(location_col_name, ""),
                            "date": row.get(date_col_name, ""),
                            "link": link_val,
                            "column": check_col_name,
                            "value": cell_val,
                            "row": row.get("#") + 1,
                        }
                        res.append(issue_item)

        if res:
            logger.warn(f"× Issues found: {len(res)}")
            if verbose:
                for item in res:
                    logger.file(item, indent=2)
            logger.mesg(self.format_check_res(res), indent=2)
        else:
            logger.okay(f"✓ All items are good!")

        self.dump_check_res(res)

        return res


class ExcelPackager:
    """Pack multiple Excel files (per week) from ExcelMerger (per day) into one file."""

    def __init__(self, date_str: str = None, past_days: int = 7):
        self.date_str = get_date_str(date_str)
        self.past_days = past_days
        self.init_dates()
        self.init_package_path()
        self.init_xlsx_paths()
        self.init_workbook()

    def init_dates(self) -> list[str]:
        """Get past N days by date_str"""
        start_date = str_to_t(self.date_str)
        self.dates: list[datetime] = []
        self.date_strs: list[str] = []
        for i in range(self.past_days, 0, -1):
            date = start_date - timedelta(days=i - 1)
            date_str = date.strftime("%Y-%m-%d")
            self.dates.append(date)
            self.date_strs.append(date_str)
        return self.date_strs

    def init_package_path(self):
        self.output_root = DATA_ROOT / "output"
        self.package_root = DATA_ROOT / "package"
        date_str_beg = self.date_strs[0].replace("-", "")
        date_str_end = self.date_strs[-1].replace("-", "")
        date_mark = f"{date_str_beg}_{date_str_end}"
        date_week = self.dates[0].isocalendar().week
        self.package_path = self.package_root / f"sku_ww{date_week}_{date_mark}.xlsx"

    def init_xlsx_paths(self) -> list[Path]:
        """Get all xlsx paths by date_strs"""
        self.xlsx_paths = []
        for date_str in self.date_strs:
            xlsx_path = self.output_root / f"{date_str}" / f"sku_{date_str}.xlsx"
            self.xlsx_paths.append(xlsx_path)
        return self.xlsx_paths

    def init_workbook(self):
        self.workbook = openpyxl.Workbook()
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

    def write_df_to_sheet(self, df: pd.DataFrame, date_str: str):
        """Write dataframe to new sheet in workbook"""
        sheet_name = f"{date_str}"
        sheet = self.workbook.create_sheet(title=sheet_name)
        # write headers
        for col_idx, column in enumerate(df.columns, 1):
            sheet.cell(row=1, column=col_idx, value=column)

        # write data
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

    def save_workbook(self):
        logger.note(f"> Save packaged xlsx to:")
        if not self.package_path.parent.exists():
            self.package_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(self.package_path)
        logger.okay(f"  * {self.package_path}")

    def package(self):
        logger.note(f"> Packaging xlsx files for:")
        for date_str, xlsx_path in zip(self.date_strs, self.xlsx_paths):
            if not xlsx_path.exists():
                logger.warn(f"  × {xlsx_path}")
                continue
            logger.file(f"  * {xlsx_path}")
            df = read_df_from_xlsx(xlsx_path)
            self.write_df_to_sheet(df, date_str)
        self.save_workbook()


class ExcelMergerArgParser(argparse.ArgumentParser):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.add_argument("-d", "--date", type=str, default=None)
        self.add_argument("-m", "--merge", action="store_true")
        self.add_argument("-k", "--check", action="store_true")
        self.add_argument("-p", "--package", action="store_true")

    def parse_args(self):
        self.args, self.unknown_args = self.parse_known_args(sys.argv[1:])
        return self.args


def main(args: argparse.Namespace):
    if args.merge:
        merger = ExcelMerger(date_str=args.date)
        merger.merge()

    if args.check:
        checker = ExcelChecker(date_str=args.date)
        checker.check()

    if args.package:
        packager = ExcelPackager(date_str=args.date)
        packager.package()


if __name__ == "__main__":
    arg_parser = ExcelMergerArgParser()
    args = arg_parser.parse_args()

    main(args)

    # Case 1: Extract data from websites and save to Excel files
    # python -m web.blinkit.batcher -e
    # python -m web.zepto.batcher -e
    # python -m web.swiggy.batcher -e
    # python -m web.dmart.batcher -e

    # Case 2: Merge Excel files (daily) into one
    # python -m file.excel_merger -m

    # Case 3: Check Excel files (daily) for missing or invalid data
    # python -m file.excel_merger -k

    # Case 4: Package Excel files (weekly) into one
    # python -m file.excel_merger -p
